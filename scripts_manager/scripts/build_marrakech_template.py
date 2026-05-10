"""Generate an empty Excel template with the 3 Marrakech sheets."""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUTPUT = os.path.join(os.path.dirname(__file__), "..", "..", "exports", "template_marrakech.xlsx")

RESTAURANTS_HEADERS = [
    "Ref", "Restaurant", "Spécialité précise", "Quartier", "Adresse",
    "Latitude", "Longitude",
    "Instagram", "Menu", "Moments", "Horaires d'ouverture",
    "Tranche de prix", "Commentaire",
    "Rooftop ?", "Dans un hôtel ?", "Festif ?",
    "Téléphone", "Site web", "Lien de réservation", "Lien Google",
]

HOTELS_HEADERS = [
    "Ref", "Hôtel / Établissement", "Quartier", "Adresse",
    "Latitude", "Longitude",
    "Catégorie", "Prix moyen", "Étoiles",
    "Instagram", "Description",
    "Piscine", "Spa", "Restaurant", "Salle de sport",
    "Fourchette (Basse / Haute saison)", "Note Google",
    "Téléphone", "Site web", "Lien de réservation",
]

DAYPASS_HEADERS = [
    "Nom", "Tag", "Quartier", "Adresse",
    "Latitude", "Longitude",
    "Instagram", "Tranche de prix", "Lien réservation",
    "Formules et prix", "Notes",
    "Téléphone", "Site web",
]

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2A2A2A", end_color="2A2A2A", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center")


def style_header_row(ws):
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    ws.row_dimensions[1].height = 22
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 4, 14), 38)
    ws.freeze_panes = "A2"


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Restaurants"
    ws.append(RESTAURANTS_HEADERS)
    style_header_row(ws)

    ws_h = wb.create_sheet("Hôtels")
    ws_h.append(HOTELS_HEADERS)
    style_header_row(ws_h)

    ws_d = wb.create_sheet("Daypass")
    ws_d.append(DAYPASS_HEADERS)
    style_header_row(ws_d)

    out_path = os.path.abspath(OUTPUT)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"Template created: {out_path}")


if __name__ == "__main__":
    main()
