#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Synchronise les attributs subscriber RevenueCat avec les données Firestore.

Pour chaque utilisateur premium dans l'Excel RC, on :
1. Matche le app_user_id (SHA256) avec un user Firestore
2. Pousse les attributs : $displayName, $email, $phoneNumber, auth_method, firebase_uid

Usage:
  python scripts/sync_revenuecat_attributes.py                          # Prod + Excel par défaut
  python scripts/sync_revenuecat_attributes.py --env dev                # Dev
  python scripts/sync_revenuecat_attributes.py --excel path/to/file.xlsx
  python scripts/sync_revenuecat_attributes.py --dry-run                # Affiche sans envoyer
  python scripts/sync_revenuecat_attributes.py --all-firestore          # Tous les users Firestore (sans Excel)
"""

import os
import sys
import hashlib
import logging
import argparse
import time
from datetime import datetime
from pathlib import Path

import requests
import openpyxl
from google.cloud import firestore
import firebase_admin
from firebase_admin import credentials, auth as fb_auth

# Config Django
sys.path.insert(0, str(Path(__file__).parent.parent))
from config import (
    EXPORTS_DIR,
    SERVICE_ACCOUNT_PATH,
    SERVICE_ACCOUNT_PATH_DEV,
    SERVICE_ACCOUNT_PATH_PROD,
)

# Charger .env pour la clé RevenueCat
from dotenv import load_dotenv
ENV_PATH = Path(__file__).resolve().parent.parent.parent / '.env'
load_dotenv(ENV_PATH)

REVENUECAT_API_KEY = os.getenv('REVENUECAT_API_KEY', '')
REVENUECAT_API_URL = 'https://api.revenuecat.com/v1/subscribers'

# Rate limiting : RevenueCat autorise ~10 req/s
BATCH_SIZE = 10
BATCH_DELAY = 1.1


def setup_logging():
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(EXPORTS_DIR / 'sync_revenuecat_attributes.log', encoding='utf-8'),
            logging.StreamHandler(),
        ],
    )
    return logging.getLogger(__name__)


def ensure_credentials(env, logger):
    if env == 'dev':
        creds_path = SERVICE_ACCOUNT_PATH_DEV
    elif env == 'prod':
        creds_path = SERVICE_ACCOUNT_PATH_PROD
    else:
        creds_path = os.environ.get('GOOGLE_APPLICATION_CREDENTIALS') or SERVICE_ACCOUNT_PATH

    if not creds_path or not Path(creds_path).exists():
        raise FileNotFoundError(f"Fichier d'identifiants non trouvé : {creds_path}")

    os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = creds_path
    logger.info(f"Identifiants Firebase: {creds_path}")
    return creds_path


def init_firebase(logger):
    if not firebase_admin._apps:
        creds_path = os.environ.get('GOOGLE_APPLICATION_CREDENTIALS') or SERVICE_ACCOUNT_PATH
        cred = credentials.Certificate(creds_path)
        firebase_admin.initialize_app(cred)
    logger.info("Firebase Admin SDK initialisé")


def sha256_hash(value):
    """Génère un hash SHA256 identique à celui du code Flutter."""
    return hashlib.sha256(value.encode('utf-8')).hexdigest()


def read_rc_app_user_ids(excel_path, logger):
    """Lit les app_user_id depuis l'Excel RevenueCat."""
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb['Revenuecat']

    headers = [cell.value for cell in ws[1]]
    app_user_id_idx = headers.index('app_user_id')
    status_idx = headers.index('status')

    rc_users = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        app_user_id = row[app_user_id_idx]
        status = row[status_idx]
        if app_user_id:
            rc_users[app_user_id] = {'status': status}

    wb.close()
    logger.info(f"{len(rc_users)} subscribers RevenueCat lus depuis l'Excel")
    return rc_users


def load_firestore_users(logger):
    """Charge tous les users Firestore avec leurs données."""
    client = firestore.Client()
    users_ref = client.collection('users')
    docs = users_ref.stream()

    users = {}
    for doc in docs:
        data = doc.to_dict()
        data['_uid'] = doc.id
        users[doc.id] = data

    logger.info(f"{len(users)} utilisateurs Firestore chargés")
    return users


def load_firebase_auth_users(logger):
    """Charge tous les users Firebase Auth pour les numéros de téléphone."""
    logger.info("Chargement des utilisateurs Firebase Auth...")
    auth_users = {}
    for user in fb_auth.list_users().iterate_all():
        auth_users[user.uid] = {
            'phone_number': getattr(user, 'phone_number', None),
            'email': getattr(user, 'email', None),
            'display_name': getattr(user, 'display_name', None),
        }
    logger.info(f"{len(auth_users)} utilisateurs Auth chargés")
    return auth_users


def match_users(rc_users, firestore_users, auth_users, logger):
    """
    Matche les app_user_id RevenueCat (SHA256) avec les users Firestore.
    Le code Flutter fait : sha256(phoneNumber ?? uid)
    """
    # Construire un index hash → uid
    hash_to_uid = {}

    for uid, fs_data in firestore_users.items():
        auth_data = auth_users.get(uid, {})
        phone = auth_data.get('phone_number') or fs_data.get('phone')

        # Le code Flutter fait : sha256(phoneNumber ?? uid)
        if phone:
            h = sha256_hash(phone)
            hash_to_uid[h] = uid
        # Fallback: sha256(uid)
        h_uid = sha256_hash(uid)
        if h_uid not in hash_to_uid:
            hash_to_uid[h_uid] = uid

    # Matcher
    matched = {}
    unmatched = []
    for app_user_id, rc_data in rc_users.items():
        uid = hash_to_uid.get(app_user_id)
        if uid:
            matched[app_user_id] = {
                'uid': uid,
                'rc_status': rc_data['status'],
                'firestore': firestore_users.get(uid, {}),
                'auth': auth_users.get(uid, {}),
            }
        else:
            unmatched.append(app_user_id)

    logger.info(f"Matchés: {len(matched)} / {len(rc_users)} ({len(unmatched)} non matchés)")
    return matched, unmatched


def build_attributes(user_data):
    """Construit les attributs à envoyer à RevenueCat."""
    fs = user_data['firestore']
    au = user_data['auth']
    uid = user_data['uid']

    attributes = {}

    # $displayName : prenom + nom
    prenom = (fs.get('prenom') or '').strip()
    nom = (fs.get('nom') or '').strip()
    display_name = ' '.join(part for part in [prenom, nom] if part)
    if display_name:
        attributes['$displayName'] = {'value': display_name}

    # $email : Firestore d'abord, puis Auth
    email = fs.get('email') or au.get('email')
    if email:
        attributes['$email'] = {'value': email}

    # $phoneNumber : Auth d'abord, puis Firestore
    phone = au.get('phone_number') or fs.get('phone')
    if phone:
        attributes['$phoneNumber'] = {'value': phone}

    # Custom: auth_method
    auth_provider = fs.get('authProvider')
    if auth_provider:
        attributes['auth_method'] = {'value': auth_provider}
    elif au.get('phone_number'):
        attributes['auth_method'] = {'value': 'phone'}

    # Custom: firebase_uid
    attributes['firebase_uid'] = {'value': uid}

    return attributes


def push_attributes(app_user_id, attributes, session, logger, dry_run=False):
    """Pousse les attributs vers RevenueCat via l'API REST."""
    if dry_run:
        name = attributes.get('$displayName', {}).get('value', '?')
        logger.info(f"  [DRY-RUN] {app_user_id[:16]}... → {name} ({len(attributes)} attributs)")
        return True

    url = f"{REVENUECAT_API_URL}/{app_user_id}/attributes"
    headers = {
        'Authorization': f'Bearer {REVENUECAT_API_KEY}',
        'Content-Type': 'application/json',
    }
    body = {'attributes': attributes}

    try:
        resp = session.post(url, json=body, headers=headers, timeout=10)
        if resp.status_code in (200, 201):
            return True
        else:
            logger.warning(f"  Erreur {resp.status_code} pour {app_user_id[:16]}...: {resp.text[:200]}")
            return False
    except requests.RequestException as e:
        logger.warning(f"  Exception pour {app_user_id[:16]}...: {e}")
        return False


def sync_all_firestore_users(firestore_users, auth_users, session, logger, dry_run=False):
    """
    Mode --all-firestore : pousse les attributs pour TOUS les users Firestore,
    en calculant leur app_user_id (SHA256) comme le fait le code Flutter.
    """
    total = len(firestore_users)
    success = 0
    skipped = 0

    items = list(firestore_users.items())
    for i, (uid, fs_data) in enumerate(items):
        auth_data = auth_users.get(uid, {})
        phone = auth_data.get('phone_number') or fs_data.get('phone')

        # Calculer le même hash que le code Flutter
        if phone:
            app_user_id = sha256_hash(phone)
        else:
            app_user_id = sha256_hash(uid)

        user_data = {
            'uid': uid,
            'firestore': fs_data,
            'auth': auth_data,
        }
        attributes = build_attributes(user_data)

        # Skip si pas d'attributs utiles (juste firebase_uid)
        if len(attributes) <= 1:
            skipped += 1
            continue

        if push_attributes(app_user_id, attributes, session, logger, dry_run):
            success += 1

        # Rate limiting
        if (i + 1) % BATCH_SIZE == 0 and i + 1 < total:
            time.sleep(BATCH_DELAY)

        if (i + 1) % 50 == 0 or i + 1 == total:
            logger.info(f"  Progression: {i + 1}/{total} (succès: {success}, skipped: {skipped})")

    return success, skipped


def main():
    parser = argparse.ArgumentParser(description="Sync attributs subscriber RevenueCat depuis Firestore")
    parser.add_argument('--env', choices=['dev', 'prod'], default='prod',
                        help="Environnement Firebase (défaut: prod)")
    parser.add_argument('--excel', type=str, default=None,
                        help="Chemin vers l'Excel RevenueCat (défaut: Users revenuecat 12.02.xlsx)")
    parser.add_argument('--dry-run', action='store_true',
                        help="Afficher sans envoyer à RevenueCat")
    parser.add_argument('--all-firestore', action='store_true',
                        help="Pousser les attributs pour TOUS les users Firestore (pas besoin d'Excel)")
    args = parser.parse_args()

    logger = setup_logging()

    if not REVENUECAT_API_KEY:
        logger.error("REVENUECAT_API_KEY non définie dans .env")
        sys.exit(1)

    logger.info(f"=== Sync attributs RevenueCat ({args.env}) ===")
    logger.info(f"Clé RC: {REVENUECAT_API_KEY[:12]}...")
    if args.dry_run:
        logger.info("MODE DRY-RUN — aucun envoi à RevenueCat")

    try:
        ensure_credentials(args.env, logger)
        init_firebase(logger)

        # Charger les données Firestore + Auth
        firestore_users = load_firestore_users(logger)
        auth_users = load_firebase_auth_users(logger)

        session = requests.Session()

        if args.all_firestore:
            # Mode: tous les users Firestore
            logger.info(f"Mode --all-firestore: sync de {len(firestore_users)} utilisateurs")
            success, skipped = sync_all_firestore_users(
                firestore_users, auth_users, session, logger, args.dry_run
            )
            logger.info(f"Terminé: {success} envoyés, {skipped} skippés")
        else:
            # Mode: depuis l'Excel RC
            excel_path = args.excel or str(Path(__file__).resolve().parent.parent.parent.parent / 'butter-final' / 'Users revenuecat 12.02.xlsx')
            if not Path(excel_path).exists():
                logger.error(f"Excel non trouvé: {excel_path}")
                sys.exit(1)

            logger.info(f"Excel: {excel_path}")
            rc_users = read_rc_app_user_ids(excel_path, logger)
            matched, unmatched = match_users(rc_users, firestore_users, auth_users, logger)

            if unmatched:
                logger.info(f"Non matchés ({len(unmatched)}): {unmatched[:5]}...")

            # Pousser les attributs
            success = 0
            errors = 0
            items = list(matched.items())

            for i, (app_user_id, user_data) in enumerate(items):
                attributes = build_attributes(user_data)

                if push_attributes(app_user_id, attributes, session, logger, args.dry_run):
                    success += 1
                else:
                    errors += 1

                # Rate limiting
                if (i + 1) % BATCH_SIZE == 0 and i + 1 < len(items):
                    time.sleep(BATCH_DELAY)

                if (i + 1) % 50 == 0 or i + 1 == len(items):
                    logger.info(f"  Progression: {i + 1}/{len(items)} (succès: {success}, erreurs: {errors})")

            logger.info(f"=== Résultat ===")
            logger.info(f"Matchés: {len(matched)}/{len(rc_users)}")
            logger.info(f"Envoyés: {success}, Erreurs: {errors}")

        session.close()

    except KeyboardInterrupt:
        logger.info("Annulé.")
    except Exception as e:
        logger.error(f"Erreur: {e}", exc_info=True)
        sys.exit(1)


if __name__ == '__main__':
    main()
