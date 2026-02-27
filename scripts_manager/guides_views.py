"""
Views pour la gestion des Guides dans le backoffice Django
Écrit directement dans Firestore (collection 'guides')
"""

import csv
import io
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from datetime import datetime
import json
import pandas as pd

from .restaurants_views import get_firestore_client
from .config import FIREBASE_BUCKET_PROD


# ==================== LISTE DES GUIDES ====================

@login_required
def guides_list(request):
    """
    Affiche la liste de tous les guides depuis Firestore
    """
    try:
        db = get_firestore_client(request)
        guides_ref = db.collection('guides')

        # Récupérer tous les guides triés par order puis name
        guides_docs = guides_ref.order_by('order').stream()

        guides = []
        premium_count = 0
        free_count = 0

        for doc in guides_docs:
            data = doc.to_dict()
            data['id'] = doc.id
            guides.append(data)

            if data.get('isPremium', False):
                premium_count += 1
            else:
                free_count += 1

        # Tri secondaire par name si même order
        guides.sort(key=lambda x: (x.get('order', 0), x.get('name', '')))

        context = {
            'guides': guides,
            'total_count': len(guides),
            'premium_count': premium_count,
            'free_count': free_count,
        }

        return render(request, 'scripts_manager/guides/list.html', context)

    except Exception as e:
        messages.error(request, f"Erreur lors du chargement des guides : {str(e)}")
        return render(request, 'scripts_manager/guides/list.html', {
            'guides': [], 'total_count': 0, 'premium_count': 0, 'free_count': 0,
        })


# ==================== DÉTAIL D'UN GUIDE ====================

@login_required
def guide_detail(request, guide_id):
    """
    Affiche le détail d'un guide
    """
    try:
        db = get_firestore_client(request)
        doc_ref = db.collection('guides').document(guide_id)
        doc = doc_ref.get()

        if not doc.exists:
            messages.error(request, f"Guide '{guide_id}' non trouvé")
            return redirect('scripts_manager:guides_list')

        guide_data = doc.to_dict()
        guide_data['id'] = doc.id

        # Récupérer les infos des restaurants référencés
        restaurant_ids = guide_data.get('restaurantIds', [])
        restaurants = []

        if restaurant_ids:
            for resto_id in restaurant_ids:
                resto_ref = db.collection('restaurants').document(resto_id)
                resto_doc = resto_ref.get()
                if resto_doc.exists:
                    resto_data = resto_doc.to_dict()
                    resto_data['id'] = resto_doc.id
                    restaurants.append(resto_data)
                else:
                    restaurants.append({
                        'id': resto_id,
                        'name': f'⚠️ Restaurant {resto_id} non trouvé',
                        'missing': True
                    })

        context = {
            'guide': guide_data,
            'restaurants': restaurants,
            'firebase_bucket': FIREBASE_BUCKET_PROD,
        }

        return render(request, 'scripts_manager/guides/detail.html', context)

    except Exception as e:
        messages.error(request, f"Erreur lors du chargement du guide : {str(e)}")
        return redirect('scripts_manager:guides_list')


# ==================== CRÉER UN GUIDE ====================

@login_required
def guide_create(request):
    """
    Formulaire de création d'un guide
    """
    # Helper : charger tous les restaurants pour le sélecteur
    def _load_all_restaurants(db_client):
        restaurants = []
        for rdoc in db_client.collection('restaurants').order_by('name').stream():
            rdata = rdoc.to_dict()
            restaurants.append({
                'id': rdoc.id,
                'name': rdata.get('name', rdoc.id),
                'cuisine': rdata.get('cuisine', ''),
                'arrondissement': rdata.get('arrondissement', ''),
                'city': rdata.get('city', 'Paris'),
            })
        return restaurants

    if request.method == 'POST':
        try:
            db = get_firestore_client(request)
            all_restaurants = _load_all_restaurants(db)

            # Récupérer les données du formulaire
            guide_id = request.POST.get('id', '').strip().upper()
            name = request.POST.get('name', '').strip()
            description = request.POST.get('description', '').strip()
            cover_image_ref = request.POST.get('coverImageRef', '').strip()
            restaurant_ids_raw = request.POST.get('restaurantIds', '').strip()
            city = request.POST.get('city', 'Paris').strip()
            order = int(request.POST.get('order', 0))
            is_premium = request.POST.get('isPremium') == 'on'
            is_featured = request.POST.get('isFeatured') == 'on'

            # Validation
            if not guide_id:
                messages.error(request, "L'ID du guide est requis")
                return render(request, 'scripts_manager/guides/form.html', {
                    'form_data': request.POST,
                    'mode': 'create',
                    'firebase_bucket': FIREBASE_BUCKET_PROD,
                    'all_restaurants': all_restaurants,
                })

            if not name:
                messages.error(request, "Le nom du guide est requis")
                return render(request, 'scripts_manager/guides/form.html', {
                    'form_data': request.POST,
                    'mode': 'create',
                    'firebase_bucket': FIREBASE_BUCKET_PROD,
                    'all_restaurants': all_restaurants,
                })

            # Vérifier que l'ID n'existe pas déjà
            doc_ref = db.collection('guides').document(guide_id)
            if doc_ref.get().exists:
                messages.error(request, f"Un guide avec l'ID '{guide_id}' existe déjà")
                return render(request, 'scripts_manager/guides/form.html', {
                    'form_data': request.POST,
                    'mode': 'create',
                    'firebase_bucket': FIREBASE_BUCKET_PROD,
                    'all_restaurants': all_restaurants,
                })

            # Parser les IDs de restaurants (séparés par virgule ou espace)
            restaurant_ids = []
            if restaurant_ids_raw:
                restaurant_ids = [
                    rid.strip()
                    for rid in restaurant_ids_raw.replace(' ', ',').split(',')
                    if rid.strip()
                ]

            # Créer le document dans Firestore
            guide_data = {
                'id': guide_id,
                'name': name,
                'description': description,
                'coverImageRef': cover_image_ref,
                'restaurantIds': restaurant_ids,
                'city': city,
                'order': order,
                'isPremium': is_premium,
                'isFeatured': is_featured,
                'createdAt': datetime.utcnow(),
                'updatedAt': datetime.utcnow(),
            }

            doc_ref.set(guide_data)

            featured_label = " (Coup de coeur)" if is_featured else ""
            premium_label = " (Premium)" if is_premium else ""
            messages.success(request, f"Guide '{name}'{premium_label}{featured_label} créé avec succès !")
            return redirect('scripts_manager:guide_detail', guide_id=guide_id)

        except ValueError as e:
            messages.error(request, f"Erreur de validation : {str(e)}")
            return render(request, 'scripts_manager/guides/form.html', {
                'form_data': request.POST,
                'mode': 'create',
                'firebase_bucket': FIREBASE_BUCKET_PROD,
                'all_restaurants': all_restaurants,
            })
        except Exception as e:
            messages.error(request, f"Erreur lors de la création du guide : {str(e)}")
            return render(request, 'scripts_manager/guides/form.html', {
                'form_data': request.POST,
                'mode': 'create',
                'firebase_bucket': FIREBASE_BUCKET_PROD,
                'all_restaurants': [],
            })

    # GET request
    try:
        db = get_firestore_client(request)
        all_restaurants = _load_all_restaurants(db)
    except Exception:
        all_restaurants = []

    return render(request, 'scripts_manager/guides/form.html', {
        'mode': 'create',
        'firebase_bucket': FIREBASE_BUCKET_PROD,
        'all_restaurants': all_restaurants,
    })


# ==================== ÉDITER UN GUIDE ====================

@login_required
def guide_edit(request, guide_id):
    """
    Formulaire d'édition d'un guide
    """
    db = get_firestore_client(request)
    doc_ref = db.collection('guides').document(guide_id)

    # Charger tous les restaurants pour le sélecteur
    all_restaurants = []
    try:
        for rdoc in db.collection('restaurants').order_by('name').stream():
            rdata = rdoc.to_dict()
            all_restaurants.append({
                'id': rdoc.id,
                'name': rdata.get('name', rdoc.id),
                'cuisine': rdata.get('cuisine', ''),
                'arrondissement': rdata.get('arrondissement', ''),
                'city': rdata.get('city', 'Paris'),
            })
    except Exception:
        pass

    if request.method == 'POST':
        try:
            # Vérifier que le guide existe
            doc = doc_ref.get()
            if not doc.exists:
                messages.error(request, f"Guide '{guide_id}' non trouvé")
                return redirect('scripts_manager:guides_list')

            # Récupérer les données du formulaire
            name = request.POST.get('name', '').strip()
            description = request.POST.get('description', '').strip()
            cover_image_ref = request.POST.get('coverImageRef', '').strip()
            restaurant_ids_raw = request.POST.get('restaurantIds', '').strip()
            city = request.POST.get('city', 'Paris').strip()
            order = int(request.POST.get('order', 0))
            is_premium = request.POST.get('isPremium') == 'on'
            is_featured = request.POST.get('isFeatured') == 'on'

            # Validation
            if not name:
                messages.error(request, "Le nom du guide est requis")
                return render(request, 'scripts_manager/guides/form.html', {
                    'guide': doc.to_dict(),
                    'guide_id': guide_id,
                    'mode': 'edit',
                    'firebase_bucket': FIREBASE_BUCKET_PROD,
                    'all_restaurants': all_restaurants,
                })

            # Parser les IDs de restaurants
            restaurant_ids = []
            if restaurant_ids_raw:
                restaurant_ids = [
                    rid.strip()
                    for rid in restaurant_ids_raw.replace(' ', ',').split(',')
                    if rid.strip()
                ]

            # Mettre à jour le document
            update_data = {
                'name': name,
                'description': description,
                'coverImageRef': cover_image_ref,
                'restaurantIds': restaurant_ids,
                'city': city,
                'order': order,
                'isPremium': is_premium,
                'isFeatured': is_featured,
                'updatedAt': datetime.utcnow(),
            }

            doc_ref.update(update_data)

            featured_label = " (Coup de coeur)" if is_featured else ""
            premium_label = " (Premium)" if is_premium else ""
            messages.success(request, f"Guide '{name}'{premium_label}{featured_label} mis à jour avec succès !")
            return redirect('scripts_manager:guide_detail', guide_id=guide_id)

        except ValueError as e:
            messages.error(request, f"Erreur de validation : {str(e)}")
            guide_data = doc_ref.get().to_dict()
            guide_data['id'] = guide_id
            return render(request, 'scripts_manager/guides/form.html', {
                'guide': guide_data,
                'guide_id': guide_id,
                'mode': 'edit',
                'firebase_bucket': FIREBASE_BUCKET_PROD,
                'all_restaurants': all_restaurants,
            })
        except Exception as e:
            messages.error(request, f"Erreur lors de la mise à jour : {str(e)}")
            guide_data = doc_ref.get().to_dict()
            guide_data['id'] = guide_id
            return render(request, 'scripts_manager/guides/form.html', {
                'guide': guide_data,
                'guide_id': guide_id,
                'mode': 'edit',
                'firebase_bucket': FIREBASE_BUCKET_PROD,
                'all_restaurants': all_restaurants,
            })

    # GET request
    try:
        doc = doc_ref.get()
        if not doc.exists:
            messages.error(request, f"Guide '{guide_id}' non trouvé")
            return redirect('scripts_manager:guides_list')

        guide_data = doc.to_dict()
        guide_data['id'] = guide_id

        # Formater restaurantIds pour l'affichage (comma-separated)
        if 'restaurantIds' in guide_data and isinstance(guide_data['restaurantIds'], list):
            guide_data['restaurantIds_display'] = ', '.join(guide_data['restaurantIds'])

        context = {
            'guide': guide_data,
            'guide_id': guide_id,
            'mode': 'edit',
            'firebase_bucket': FIREBASE_BUCKET_PROD,
            'all_restaurants': all_restaurants,
        }

        return render(request, 'scripts_manager/guides/form.html', context)

    except Exception as e:
        messages.error(request, f"Erreur lors du chargement du guide : {str(e)}")
        return redirect('scripts_manager:guides_list')


# ==================== SUPPRIMER UN GUIDE ====================

@login_required
@require_http_methods(["POST"])
def guide_delete(request, guide_id):
    """
    Supprime un guide de Firestore
    """
    try:
        db = get_firestore_client(request)
        doc_ref = db.collection('guides').document(guide_id)

        # Vérifier que le guide existe
        doc = doc_ref.get()
        if not doc.exists:
            messages.error(request, f"Guide '{guide_id}' non trouvé")
            return redirect('scripts_manager:guides_list')

        guide_name = doc.to_dict().get('name', guide_id)

        # Supprimer le document
        doc_ref.delete()

        messages.success(request, f"Guide '{guide_name}' supprimé avec succès")
        return redirect('scripts_manager:guides_list')

    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression : {str(e)}")
        return redirect('scripts_manager:guides_list')


# ==================== EXPORT JSON ====================

@login_required
def guide_get_json(request, guide_id):
    """
    Retourne le JSON d'un guide (pour debug / export)
    """
    try:
        db = get_firestore_client(request)
        doc_ref = db.collection('guides').document(guide_id)
        doc = doc_ref.get()

        if not doc.exists:
            return JsonResponse({'error': 'Guide non trouvé'}, status=404)

        guide_data = doc.to_dict()
        guide_data['id'] = doc.id

        # Convertir les timestamps en strings
        if 'createdAt' in guide_data:
            guide_data['createdAt'] = guide_data['createdAt'].isoformat() if hasattr(guide_data['createdAt'], 'isoformat') else str(guide_data['createdAt'])
        if 'updatedAt' in guide_data:
            guide_data['updatedAt'] = guide_data['updatedAt'].isoformat() if hasattr(guide_data['updatedAt'], 'isoformat') else str(guide_data['updatedAt'])

        return JsonResponse(guide_data, json_dumps_params={'indent': 2})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ==================== IMPORT EXCEL ====================

def _cell_str(value):
    """Convertit une cellule Excel en chaîne (None, int, float → str)."""
    if value is None:
        return ''
    return str(value).strip()


@login_required
def guides_import_csv(request):
    """
    Import de guides depuis un fichier Excel (.xlsx) ou CSV (.csv).
    Format attendu (1ère ligne = en-têtes) : Nom, Ref, Description, Restaurants, Photo couverture
    """
    if request.method == 'POST':
        try:
            uploaded = request.FILES.get('excel_file') or request.FILES.get('csv_file')
            if not uploaded:
                messages.error(request, "Aucun fichier fourni.")
                return redirect('scripts_manager:guides_list')

            name_lower = (uploaded.name or '').lower()
            is_csv = name_lower.endswith('.csv')

            if not (name_lower.endswith('.xlsx') or name_lower.endswith('.xls') or is_csv):
                messages.error(request, "Le fichier doit être un Excel (.xlsx, .xls) ou CSV (.csv).")
                return redirect('scripts_manager:guides_list')

            if is_csv:
                # Lire le CSV avec pandas (multi-encodage)
                import tempfile
                import os
                suffix = '.csv'
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                    for chunk in uploaded.chunks():
                        tmp.write(chunk)
                    tmp_path = tmp.name
                try:
                    df = None
                    for enc in ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
                        try:
                            df = pd.read_csv(tmp_path, encoding=enc, sep=None, engine='python')
                            break
                        except UnicodeDecodeError:
                            continue
                    if df is None:
                        messages.error(request, "Encodage CSV non supporté.")
                        return redirect('scripts_manager:guides_list')
                    # Convertir en liste de tuples (header + rows)
                    rows = [tuple(df.columns)] + [tuple(row) for row in df.values]
                finally:
                    os.unlink(tmp_path)
            else:
                from openpyxl import load_workbook
                wb = load_workbook(filename=uploaded, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                wb.close()

            if not rows:
                messages.error(request, "Le fichier est vide.")
                return redirect('scripts_manager:guides_list')

            # 1ère ligne = en-têtes
            headers = [_cell_str(c) for c in rows[0]]
            col_map = {}
            for i, h in enumerate(headers):
                h_lower = (h or '').lower().strip()
                if not h_lower:
                    continue
                if h_lower == 'nom' and 'Nom' not in col_map:
                    col_map['Nom'] = i
                elif h_lower == 'ref' and 'Ref' not in col_map:
                    col_map['Ref'] = i
                elif 'description' in h_lower and 'Description' not in col_map:
                    col_map['Description'] = i
                elif 'restaurant' in h_lower and 'Restaurants' not in col_map:
                    col_map['Restaurants'] = i
                elif ('photo' in h_lower and 'couverture' in h_lower) and 'Photo couverture' not in col_map:
                    col_map['Photo couverture'] = i
                elif h_lower == 'premium' and 'Premium' not in col_map:
                    col_map['Premium'] = i
                elif ('coup' in h_lower or 'featured' in h_lower) and 'Featured' not in col_map:
                    col_map['Featured'] = i

            def get_cell(row, key):
                idx = col_map.get(key)
                if idx is None or idx >= len(row):
                    return ''
                return _cell_str(row[idx])

            db = get_firestore_client(request)
            imported_count = 0
            errors = []

            for row_num, row in enumerate(rows[1:], start=2):
                try:
                    row = list(row) if row else []
                    name = get_cell(row, 'Nom')
                    guide_id = get_cell(row, 'Ref').upper()
                    description = get_cell(row, 'Description')
                    restaurants_raw = get_cell(row, 'Restaurants')
                    cover_image_ref = get_cell(row, 'Photo couverture')
                    premium_raw = get_cell(row, 'Premium').lower()
                    is_premium = premium_raw in ['oui', 'yes', 'true', '1', 'premium']
                    featured_raw = get_cell(row, 'Featured').lower()
                    is_featured = featured_raw in ['oui', 'yes', 'true', '1', 'featured']

                    if not guide_id or not name:
                        errors.append(f"Ligne {row_num} : ID (Ref) ou nom manquant")
                        continue

                    restaurant_ids = [
                        rid.strip()
                        for rid in restaurants_raw.replace(' ', ',').split(',')
                        if rid.strip()
                    ]

                    doc_ref = db.collection('guides').document(guide_id)
                    guide_data = {
                        'id': guide_id,
                        'name': name,
                        'description': description,
                        'coverImageRef': cover_image_ref,
                        'restaurantIds': restaurant_ids,
                        'isPremium': is_premium,
                        'isFeatured': is_featured,
                        'order': imported_count,
                        'updatedAt': datetime.utcnow(),
                    }
                    if not doc_ref.get().exists:
                        guide_data['createdAt'] = datetime.utcnow()

                    doc_ref.set(guide_data, merge=True)
                    imported_count += 1

                except Exception as e:
                    errors.append(f"Ligne {row_num} : {str(e)}")

            if imported_count > 0:
                messages.success(request, f"{imported_count} guide(s) importé(s) avec succès")

            if errors:
                error_msg = "Erreurs lors de l'import :\n" + "\n".join(errors[:10])
                if len(errors) > 10:
                    error_msg += f"\n... et {len(errors) - 10} autres erreurs"
                messages.warning(request, error_msg)

            return redirect('scripts_manager:guides_list')

        except Exception as e:
            messages.error(request, f"Erreur lors de l'import : {str(e)}")
            return redirect('scripts_manager:guides_list')

    return render(request, 'scripts_manager/guides/import.html')


# ==================== EXPORT CSV / EXCEL ====================

@login_required
def guides_export(request):
    """Exporte tous les guides en CSV ou Excel."""
    fmt = request.GET.get('format', 'csv')

    try:
        db = get_firestore_client(request)
        guides_docs = db.collection('guides').order_by('order').stream()

        guides = []
        for doc in guides_docs:
            data = doc.to_dict()
            data['id'] = doc.id
            guides.append(data)

        guides.sort(key=lambda x: (x.get('order', 0), x.get('name', '')))

        headers = ['Nom', 'Ref', 'Description', 'Restaurants', 'Photo couverture', 'Premium', 'Coup de coeur']
        rows = []
        for g in guides:
            restaurant_ids = g.get('restaurantIds', [])
            rows.append({
                'Nom': g.get('name', ''),
                'Ref': g.get('id', ''),
                'Description': g.get('description', '') or '',
                'Restaurants': ', '.join(restaurant_ids) if restaurant_ids else '',
                'Photo couverture': g.get('coverImageRef', '') or '',
                'Premium': 'oui' if g.get('isPremium') else 'non',
                'Coup de coeur': 'oui' if g.get('isFeatured') else 'non',
            })

        if fmt == 'xlsx':
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Guides'
            ws.append(headers)
            for row in rows:
                ws.append([row.get(h, '') for h in headers])
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="guides.xlsx"'
            return response
        else:
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="guides.csv"'
            response.write('\ufeff')
            writer = csv.DictWriter(response, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)
            return response

    except Exception as e:
        messages.error(request, f"Erreur lors de l'export : {str(e)}")
        return redirect('scripts_manager:guides_list')
